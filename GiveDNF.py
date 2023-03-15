import wx
import os
import openpyxl
import datetime
import Model
import Utils
from ExportGrid import ExportGrid

def GiveDNF( bibs, keepCategoryTheSame=True ):
	race = Model.race
	if not race:
		return

	#Get the reference Excel sheet.
	#excelLink = getattr( race, 'excelLink', None )
	#tNow = datetime.datetime.now()
	#if excelLink:
		#Open the Excel file.  This may throw an exception.
		#wb = openpyxl.load_workbook( filename=excelLink.fileName )
		
		#Make a backup copy of the Excel workbook.  This may throw an exception.
		#fnameBackup = os.path.splitext(excelLink.fileName)[0] + '-backup-' + tNow.strftime('%Y-%m-%d-%H%M%S') + '.xlsx'
		#wb.save( fnameBackup )
		
		#Get the sheet.  This may throw an exception.
		#ws = wb[excelLink.sheetName]
		#bibCol = excelLink.fieldCol['Bib#'] + 1		# Add 1 because openlyxl indexes from 1.
	#else:
		#wb = ws = bibCol = None

	#Fix up all the bib numbers in Excel.
	#if ws:
		#bibDict = { bibOld:bibNew for bibOld,bibNew in bibs }
		#foundHeader = False
		#for row in range(1, ws.max_row+1):
			#if not foundHeader:
				#foundHeader = (sum( int(bool(ws.cell(row, col).value)) for col in range(1, ws.max_column+1) ) > 4)
				#continue
			#cell = ws.cell( row, bibCol )
			#value = cell.value
			#if isinstance(value, (float, int)) and value == int(value):
				#value = int( value )
				#if value in bibDict:
					#cell.value = bibDict[cell.value]

	#Save the modified Excel file.  This may throw an exception.
	#if wb:
		#wb.save( excelLink.fileName )	
	
	#-------------------------------------------------------------------
	#Done with the Excel file.
	#-------------------------------------------------------------------
	
	#Get all categories and their number sets in search order.
	#CatWave = Model.Category.CatWave
	#CatComponent = Model.Category.CatComponent
	#CatCustom = Model.Category.CatCustom

	#categorySets, customCategorySets = [], []
	#categories = race.getCategories( startWaveOnly=False )
	#for i, c in enumerate(categories):
		#if c.catType == CatCustom:
			#customCategorySets.append( [c, c.getMatchSet()] )
			#continue
		#if c.catType == CatWave:
			#If this Wave has components, get the components, not the wave.
			#The bib numbers of the wave are dependent on the components.
			#try:
				#cNext = categories[i+1]
				#if cNext.catType == CatComponent:
					#continue
			#except IndexError:
				#pass
		#categorySets.append( [c, c.getMatchSet()] )

	#Routine to find numbers sets by category.
	#def getCategorySet( num, categorySets ):
		#for c, cSet in categorySets:
			#if num in cSet:
				#return c, cSet
		#return None, None
	
	#for bibOld, bibNew in bibs:
		#Ensure that the change will keep the rider in the same category.
		#This has to be done separately for the regular categories and the custom categories.
		#if keepCategoryTheSame:
			#for cs in [categorySets, customCategorySets]:
				#while True:
					#cOld, cSetOld = getCategorySet( bibOld, cs )
					#cNew, cSetNew = getCategorySet( bibNew, cs )
					#if not (cOld and cOld != cNew):
						#break
					
					#cSetOld.add( bibNew )
					#if cSetNew:
						#cSetNew.discard( bibNew )

		#riderNew = race.riders.get( bibNew, None )
		#if riderNew:
			#riderNew.clearCache()
		#riderOld = race.riders.get( bibOld, None )
		#if riderOld:
			#riderOld.clearCache()
		
		#if riderNew and riderOld:
			#Copy the attributes of the new to the old to preserve any TT start-times.
			#for a in ('num', 'times', 'status', 'tStatus', 'autocorrectLaps'):
				#setattr( riderOld, a, getattr(riderNew, a) )
			#If the new bib exists and has a firstTime, this is actually the first time entered.
			#Add it as a recorded time.
			#if race.isTimeTrial and riderNew.firstTime and riderNew.firstTime > riderOld.firstTime:
				#riderOld.times.append( riderNew.firstTime )
				#riderOld.times.sort()
			#del race.riders[bibOld]
			#race.riders[bibNew] = riderOld
		#elif riderNew and not riderOld:
			#pass								# No pre-existing bibOld data.  The rider attributes will be picked up from the Excel sheet.
		#elif not riderNew and riderOld:
			#del race.riders[bibOld]				# No data on bibNew.  Just change the rider's bib number.
			#riderOld.num = bibNew
			#race.riders[bibNew] = riderOld
	
	#Reset the race category sets.
	#for c, cSet in categorySets:
		#c.setFromSet( cSet )
	#race.adjustAllCategoryWaveNumbers()	# Reset the compressed ranges of the waves.
	#race.setChanged()

class GiveDNFDialog( wx.Dialog ):
	excludeStatuses = ['Finisher', 'DQ', 'DNS' ]
	
	def __init__( self, parent, id = wx.ID_ANY ):
		super().__init__(parent, id, title=_('Give non-finishers a finish time'))
		self.leaderTime ='0:00:00.00'
		self.useLastFinisherTime = '0:00:00.00'
		self.SetBackgroundColour( wx.WHITE )
		
		vs = wx.BoxSizer( wx.VERTICAL )

		title = wx.StaticText( self, label=_("Gives riders who did not finish a virtual finish time.") )
		explain = []
		explain.append( wx.StaticText( self, label=_("Be careful!  There is no Undo.") ) )
		explain.append( wx.StaticText( self, label=_("This modifies the final lap time of riders who did not finish the race and sets their status to 'Finisher'.") ) )
		explain.append( wx.StaticText( self, label=_("Their average speed and ranked position will therefore reflect the number of laps that they were") ) )
		explain.append( wx.StaticText( self, label=_("able to complete successfully.") ) )
		
		fgs = wx.FlexGridSizer(vgap=4, hgap=4, rows=0, cols=2)
		fgs.AddGrowableCol( 1, 1 )
		
		GetTranslation = _
		self.statusName = wx.StaticText( self, label = '{} '.format(_('Change finish time of riders with status:')) )
		fgs.Add( self.statusName, 1, flag=wx.TOP|wx.ALIGN_LEFT )
		self.statusOption = wx.Choice( self, choices=[GetTranslation(n) for n in Model.Rider.statusNames if n not in self.excludeStatuses] )
		fgs.Add( self.statusOption, 1, flag=wx.TOP|wx.ALIGN_LEFT )
		self.Bind(wx.EVT_CHOICE, self.onStatusChanged, self.statusOption)
		
		self.useLeaderTime = wx.RadioButton( self, label=_("Winner's time"), style = wx.RB_GROUP )
		self.useLastFinisherTime = wx.RadioButton( self, label=_("Last finisher's time") )
		self.useNominalTime = wx.RadioButton( self, label=_("Nominal race time") )
		self.useCustomTime = wx.RadioButton( self, label=_("Custom") )
		self.Bind( wx.EVT_RADIOBUTTON,self.onSelectTime ) 
		fgs.Add( wx.StaticText( self, label = _("To:")), flag=wx.ALIGN_RIGHT|wx.ALIGN_CENTRE_VERTICAL )
		fgs.Add( self.useLeaderTime, 1, flag=wx.TOP|wx.ALIGN_LEFT )
		fgs.AddSpacer( 16 )
		fgs.Add( self.useLastFinisherTime, 1, flag=wx.TOP|wx.ALIGN_LEFT )
		fgs.AddSpacer( 16 )
		fgs.Add( self.useNominalTime, 1, flag=wx.TOP|wx.ALIGN_LEFT )
		fgs.AddSpacer( 16 )
		fgs.Add( self.useCustomTime, 1, flag=wx.TOP|wx.ALIGN_LEFT )
		
		fgs.Add( wx.StaticText( self, label = _("Time:")), flag=wx.ALIGN_RIGHT|wx.ALIGN_CENTRE_VERTICAL )
		self.timeToUse = wx.TextCtrl( self, size=(256,-1), style=wx.TE_PROCESS_ENTER, value='12345' )
		fgs.Add (self.timeToUse, 1, flag=wx.TOP|wx.ALIGN_LEFT )

		vs.Add( title, flag=wx.ALL, border=4 )
		vs.AddSpacer( 8 )
		for e in explain:
			vs.Add( e, flag=wx.LEFT|wx.RIGHT, border=4 )
		vs.AddSpacer( 8 )
		
		vs.Add(fgs, flag=wx.ALL, border=4 )
			
		btnSizer = self.CreateStdDialogButtonSizer( wx.OK|wx.CANCEL )
		self.Bind( wx.EVT_BUTTON, self.onOK, id=wx.ID_OK )
		if btnSizer:
			vs.Add( btnSizer, flag = wx.ALL|wx.EXPAND, border=4 )
			
		self.SetSizerAndFit( vs )
		self.refresh()
		
	def onSelectTime( self, event ):
		if self.useLeaderTime.GetValue():
			self.timeToUse.SetValue(Utils.formatTime(self.leaderTime, highPrecision=True, forceHours=True))
		elif self.useLastFinisherTime.GetValue():
			self.timeToUse.SetValue(Utils.formatTime(self.lastFinisherTime, highPrecision=True, forceHours=True))
		elif self.useNominalTime.GetValue():
			self.timeToUse.SetValue(Utils.formatTime(self.race.minutes*60, highPrecision=True, forceHours=True))
		elif self.useCustomTime.GetValue():
			self.timeToUse.SetValue('0:00:00.00')
		return
	
	def onStatusChanged( self, event ):
		pass
	
	def refresh( self ):
		race = Model.race
		if not race:
			return
		self.race = race
		self.lastFinisherTime, self.lastFinisher = race.getLastKnownTimeRider()
		self.leader, self.leaderTime, lap = race.getLeaderTimeLap()
		self.useNominalTime.SetValue(True)
		self.timeToUse.SetValue(Utils.formatTime(self.race.minutes*60, highPrecision=True, forceHours=True))

	def onOK( self, event ):
		#bibs = {}
		#bibsOld = set()
		#bibsNew = set()
		#for row in range(self.grid.GetNumberRows()):
			#try:
				#bibOld = int(self.grid.GetCellValue(row, 0))
				#if not bibOld:
					#continue
			#except ValueError:
				#continue
			#try:
				#bibNew = int(self.grid.GetCellValue(row, 1))
				#if not bibNew:
					#continue
			#except ValueError:
				#continue
				
			#bibs[bibOld] = bibNew
			#bibsOld.add( bibOld )
			#bibsNew.add( bibNew )

		#def clearGrid():
			#for row in range(self.grid.GetNumberRows()):
				#self.grid.SetCellValue( row, 0, '' )
				#self.grid.SetCellValue( row, 1, '' )
			
		#if not bibs:
			#clearGrid()
			#self.EndModal( wx.ID_OK )
			
		#bibsOldNew = bibsOld & bibsNew
		#if bibsOldNew:
			#message = '{}\n\n\t{}\n\n{}'.format(
				#_('The following Bibs are in both the Old and New columns:'),
				#','.join( str(bib) for bib in sorted(bibsOldNew) ),
				#_('Please correct and retry.'),
			#)
			#with wx.MessageDialog(self, message, _('Reissue Bibs Error'), ) as dlg:
				#dlg.ShowModal()
			#Utils.refresh()
			#return				
		
		#bibs = list( bibs.items() )
		
		#try:
			#ReissueBibs( bibs )
		#except Exception as e:
			#with wx.MessageDialog(self, '{}: {}'.format(_('Error'), e), _('Reissue Bibs Error'), ) as dlg:
				#dlg.ShowModal()
			#Utils.refresh()
			#return	
			
		#clearGrid()
		Utils.refresh()
		self.EndModal( wx.ID_OK )
		
if __name__ == '__main__':
	Utils.disable_stdout_buffering()
	app = wx.App(False)
	mainWin = wx.Frame(None,title="CrossMan", size=(600,600))
	giveDNF = GiveDNFDialog(mainWin)
	mainWin.Show()
	giveDNF.ShowModal()
	app.MainLoop()	
